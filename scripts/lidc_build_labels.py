"""
Builds a per-nodule label table from the LIDC-IDRI radiologist annotations.

This is the half of the retraining pipeline that needs no pixel data. The XML
annotations are 170 MB; the images are 128 GB. Doing the labels first means the
download can be a targeted subset rather than the whole collection, and it means
the label logic is settled and inspectable before a single byte of imaging is
fetched.

    python scripts/lidc_build_labels.py "TCIA Folder" dataset/lidc-labels.csv

── How a malignancy label is derived ──────────────────────────────────────

LIDC does not ship a cancer / no-cancer flag. Up to four thoracic radiologists
independently mark each nodule >=3 mm and rate malignancy 1-5:

    1  highly unlikely      2  moderately unlikely   3  indeterminate
    4  moderately suspicious                         5  highly suspicious

Two problems have to be solved before that becomes a training label.

**Nodule identity.** Each radiologist assigns their own `noduleID`, so the same
physical nodule carries up to four different identifiers. Readings are therefore
clustered by the 3-D position of their ROI centroids: readings whose centroids
fall within `--match-mm` of each other are treated as one nodule. This is the
standard approach and the reason `pylidc` exists; it is implemented here
directly to avoid a dependency for one function.

**Disagreement.** The median across readers is used, which is the convention in
the literature. A nodule whose median is exactly 3 is *excluded* rather than
pushed to one side — "the radiologists could not agree, or agreed it was
indeterminate" is not a training signal for a binary classifier, and forcing it
into one class would put roughly 38% of the data on the wrong side of a
threshold nobody drew.

Also recorded, and worth using later: `agreement`, the fraction of readers whose
individual rating falls on the same side of 3 as the median. A nodule labelled
malignant by four readers out of four is not the same evidence as one labelled
malignant by two out of four, and a training run may reasonably weight or filter
on it.

── What this deliberately does not do ─────────────────────────────────────

It does not use the Patient Diagnoses file. That file carries genuine
diagnosis-level truth — biopsy or follow-up confirmed — which is stronger
evidence than a radiologist's visual impression. It also covers only a limited
subset of the collection, and TCIA states the project ended without collecting
more. Building on it would give a small, high-quality set; building on the
ratings gives a large, weaker-labelled one. The ratings are used here because
the immediate problem is that the model has never seen a real CT at all, and
that is a distribution problem before it is a label-quality problem.

Anything published from a model trained this way must say which it used. A
malignancy *rating* is a radiologist's impression from the image; a *diagnosis*
is what the patient turned out to have. They are not interchangeable and the
model card must not blur them.
"""
from __future__ import annotations

import argparse
import collections
import csv
import glob
import os
import statistics
import sys
import xml.etree.ElementTree as ET

NS = "{http://www.nih.gov}"


def _text(node, tag):
    found = node.find(f"{NS}{tag}")
    return (found.text or "").strip() if found is not None and found.text else ""


def parse_reading_sessions(path):
    """Every nodule reading in one XML: (series_uid, malignancy, centroid, slices)."""
    try:
        root = ET.parse(path).getroot()
    except ET.ParseError:
        return None, []

    header = root.find(f"{NS}ResponseHeader")
    series_uid = _text(header, "SeriesInstanceUid") if header is not None else ""

    readings = []
    for session in root.findall(f"{NS}readingSession"):
        for nodule in session.findall(f"{NS}unblindedReadNodule"):
            characteristics = nodule.find(f"{NS}characteristics")
            if characteristics is None:
                # No characteristics means a nodule <3 mm. LIDC does not rate
                # those, so there is no label to derive and they are skipped
                # rather than treated as benign — an unrated nodule is not a
                # negative finding.
                continue

            raw = _text(characteristics, "malignancy")
            if not raw.isdigit():
                continue
            malignancy = int(raw)
            if not 1 <= malignancy <= 5:
                continue

            xs, ys, zs, sops = [], [], [], []
            for roi in nodule.findall(f"{NS}roi"):
                z = _text(roi, "imageZposition")
                sop = _text(roi, "imageSOP_UID")
                edges = roi.findall(f"{NS}edgeMap")
                if not edges:
                    continue
                ex = [int(_text(e, "xCoord")) for e in edges if _text(e, "xCoord").lstrip("-").isdigit()]
                ey = [int(_text(e, "yCoord")) for e in edges if _text(e, "yCoord").lstrip("-").isdigit()]
                if not ex or not ey:
                    continue
                xs.append(sum(ex) / len(ex))
                ys.append(sum(ey) / len(ey))
                if z:
                    try:
                        zs.append(float(z))
                    except ValueError:
                        pass
                if sop:
                    sops.append((float(z) if z else 0.0, sop, sum(ex) / len(ex), sum(ey) / len(ey)))

            if not xs or not zs:
                continue

            readings.append({
                "malignancy": malignancy,
                "cx": sum(xs) / len(xs),
                "cy": sum(ys) / len(ys),
                "cz": sum(zs) / len(zs),
                "slices": sops,
            })

    return series_uid, readings


def cluster(readings, match_mm):
    """Groups readings of the same physical nodule.

    Greedy single-link clustering on centroid distance. The in-plane coordinates
    are pixels and z is millimetres, which is not dimensionally consistent — but
    LIDC in-plane spacing is close to 0.7 mm and nodules are separated by far
    more than the error this introduces, so a single threshold works and a
    correct conversion would need the DICOM we do not have yet.
    """
    clusters = []
    for reading in readings:
        placed = False
        for group in clusters:
            head = group[0]
            distance = (
                (reading["cx"] - head["cx"]) ** 2
                + (reading["cy"] - head["cy"]) ** 2
                + (reading["cz"] - head["cz"]) ** 2
            ) ** 0.5
            if distance <= match_mm:
                group.append(reading)
                placed = True
                break
        if not placed:
            clusters.append([reading])
    return clusters


def load_digest(folder):
    """Series UID -> patient, size, image count. Needs openpyxl."""
    import openpyxl

    matches = glob.glob(os.path.join(folder, "*digest*.xlsx"))
    if not matches:
        return {}

    workbook = openpyxl.load_workbook(matches[0], read_only=True)
    sheet = workbook["Metadata"] if "Metadata" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    index = {name: position for position, name in enumerate(header)}

    out = {}
    for row in rows:
        if row[index["Modality"]] != "CT":
            continue
        out[row[index["Series Instance UID"]]] = {
            "patient": row[index["Patient ID"]],
            "bytes": row[index["File Size"]] or 0,
            "images": row[index["Image Count"]] or 0,
        }
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("tcia_folder")
    parser.add_argument("out_csv")
    parser.add_argument("--match-mm", type=float, default=12.0,
                        help="centroid distance below which readings are one nodule")
    args = parser.parse_args()

    digest = load_digest(args.tcia_folder)
    xml_files = glob.glob(os.path.join(args.tcia_folder, "**", "*.xml"), recursive=True)
    if not xml_files:
        print(f"no XML under {args.tcia_folder}", file=sys.stderr)
        raise SystemExit(1)

    rows = []
    unmatched_series = 0
    stats = collections.Counter()

    for path in xml_files:
        series_uid, readings = parse_reading_sessions(path)
        if not readings:
            continue
        meta = digest.get(series_uid)
        if meta is None:
            unmatched_series += 1
            continue

        for position, group in enumerate(cluster(readings, args.match_mm)):
            scores = [r["malignancy"] for r in group]
            median = statistics.median(scores)

            if median > 3:
                label, positive = "malignant", 1
            elif median < 3:
                label, positive = "benign", 0
            else:
                stats["excluded_indeterminate"] += 1
                continue

            same_side = sum(1 for s in scores if (s > 3) == (median > 3))
            best = max(group, key=lambda r: len(r["slices"]))
            centre_slice = sorted(best["slices"], key=lambda s: s[0])[len(best["slices"]) // 2] if best["slices"] else None

            rows.append({
                "patient": meta["patient"],
                "series_uid": series_uid,
                "nodule_index": position,
                "label": label,
                "positive": positive,
                "median_malignancy": median,
                "n_readers": len(scores),
                "agreement": round(same_side / len(scores), 2),
                "ratings": "|".join(str(s) for s in scores),
                "sop_uid": centre_slice[1] if centre_slice else "",
                "z": round(centre_slice[0], 2) if centre_slice else "",
                "cx": round(centre_slice[2]) if centre_slice else "",
                "cy": round(centre_slice[3]) if centre_slice else "",
                "series_bytes": meta["bytes"],
            })
            stats[label] += 1

    os.makedirs(os.path.dirname(args.out_csv) or ".", exist_ok=True)
    with open(args.out_csv, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    patients = {r["patient"] for r in rows}
    series = {r["series_uid"] for r in rows}
    confident = [r for r in rows if r["agreement"] >= 0.75]

    print(f"wrote {args.out_csv}")
    print()
    print(f"  labelled nodules      : {len(rows)}")
    print(f"    benign              : {stats['benign']}")
    print(f"    malignant           : {stats['malignant']}")
    print(f"  excluded (median = 3) : {stats['excluded_indeterminate']}")
    print(f"  >=75% reader agreement: {len(confident)}")
    print()
    print(f"  patients              : {len(patients)}")
    print(f"  CT series             : {len(series)}")
    print(f"  series in XML but not in digest: {unmatched_series}")
    total = sum({r['series_uid']: r['series_bytes'] for r in rows}.values())
    print(f"  download for ALL of these series: {total/1e9:.1f} GB")


if __name__ == "__main__":
    main()
